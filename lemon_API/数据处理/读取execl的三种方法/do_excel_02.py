# -*- coding: utf-8 -*-
# @Time :2020/6/20 14:51
# @Author : liufei
# @File :do_excel_02.PY

                                        #方法二：根据传入的坐标来获取值
#excel 中的所有数据除了数字 还是数字类型（int还是int,float还是float），其他数据格式均变为字符串类型（list,tuple,dict,均变成了字符串格式）

from openpyxl import load_workbook
class GetTestData():
    def __init__(self,filename,sheetname):
        self.filename=filename
        self.sheetname=sheetname
        self.sheet_obj=load_workbook(self.filename)[self.sheetname]  #获取一个表单对象
        self.max_row=self.sheet_obj.max_row
    def get_data(self,i,j):
        return self.sheet_obj.cell(i,j).value

if __name__ == '__main__':
    print(GetTestData('data_01-02.xlsx', 'request_data').get_data(1, 1))
