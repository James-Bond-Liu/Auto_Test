# -*- coding: utf-8 -*-
# @Time :2020/6/20 13:57
# @Author : liufei
# @File :do_excel.PY

                                        #方法一：一次性的从excel中获得所有的数据
#excel 中的所有数据除了数字 还是数字类型（int还是int,float还是float），其他数据格式均变为字符串类型（list,tuple,dict,均变成了字符串格式）

from openpyxl import load_workbook
#openpyxl处理execl行列的索引是从“1”开始的

class GetTestData():
    def __init__(self,filename,sheetname):
        self.filename=filename
        self.sheetname=sheetname

    def get_test_data(self):
        wb=load_workbook(self.filename)
        ws=wb[self.sheetname]  # 或者通过ws = wb.get_sheet_by_name(‘sheet表单名’)定位到某个表单
        print(wb.get_sheet_names())  #返回excel文件中所有的sheet表单名并保存在列表中
        test_data=[]
        for i in range(1,ws.max_row+1):
            # sub_data={}
            # sub_data['url']=ws.cell(i,1).value
            # sub_data['methond']=ws.cell(i,2).value
            # sub_data['data']=ws.cell(i,3).value
            # sub_data['expected']=ws.cell(i,4).value
            # test_data.append(sub_data)

            sub_data= {'url': ws.cell(i, 1).value,
                       'methond': ws.cell(i, 2).value,
                       'data': ws.cell(i,3).value,
                   'expected': ws.cell(i, 4).value}
            test_data.append(sub_data)
        return test_data
if __name__ == '__main__':
    print(GetTestData('data_01-02.xlsx', 'request_data').get_test_data())
