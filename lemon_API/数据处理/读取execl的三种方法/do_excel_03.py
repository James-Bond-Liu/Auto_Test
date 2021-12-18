# -*- coding: utf-8 -*-
# @Time :2020/6/20 18:27
# @Author : liufei
# @File :do_excel_03.PY

                                #提取excel中数据方法三，主要应用在excel中存在表头时，例如data_03,
from openpyxl import load_workbook

class GetRequestData():
    def __init__(self,filename,sheetname):
        self.filename=filename
        self.sheetname=sheetname

    def get_header(self):#获取excel中的表头
        header=[]
        wb = load_workbook(self.filename)
        ws = wb[self.sheetname]
        for j in range(1,ws.max_column+1):
            header.append(ws.cell(1,j).value)
        return header

    def get_test_data(self):
       # 提取数据，并将header作为key，提取的数据作为“value”
        wb=load_workbook(self.filename)
        ws=wb[self.sheetname]
        test_data=[]
        header=self.get_header() #拿到上一个函数取到的header，同一类下调用其他方法时要加self.方法名
        for i in range(2,ws.max_row+1):
            sub_data={}#一行数据存到字典中
            for j in range(1,ws.max_column+1):
                sub_data[header[j-1]]=ws.cell(i,j).value
            test_data.append(sub_data)#在列表中嵌套字典，将刚提取出来的数据追加到字典中，然后进入下一轮循环
        return test_data

if __name__ == '__main__':
    print(GetRequestData('data_03.xlsx','request_data').get_test_data())

