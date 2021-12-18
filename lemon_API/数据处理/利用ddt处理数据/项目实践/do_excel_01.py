# -*- coding: utf-8 -*-
# @Time :2020/6/20 13:57
# @Author : liufei
# @File :do_excel.PY

                                #利用配置文件来控制用例,实现用例的可配置
from openpyxl import load_workbook
from 数据处理.利用ddt处理数据.项目实践.read_config import ReadConfig

class GetTestData():
    def __init__(self,filename,sheetname):
        self.filename=filename
        self.sheetname=sheetname

    def get_test_data(self):
        mode=ReadConfig().read_config('case.config', 'MODE', 'mode')#直接从配置文件读取mode值
        wb=load_workbook(self.filename)
        ws=wb[self.sheetname]
        test_data=[]
        for i in range(1,ws.max_row+1):
            # sub_data={}
            # sub_data['url']=ws.cell(i,ws.max_row).value
            # sub_data['methond']=ws.cell(i,ws.max_row).value
            # sub_data['data']=ws.cell(i,ws.max_row).value
            # sub_data['expected']=ws.cell(i,ws.max_row).value
            # test_data.append(sub_data)

            sub_data= {'case_id':ws.cell(i, 1).value,
                'url': ws.cell(i, 2).value,
                'methond': ws.cell(i, 3).value,
                'data': ws.cell(i,4).value,
                'expected': ws.cell(i, 5).value}
            test_data.append(sub_data)

        if mode=='all':#数据的控制，根据mode的值进行判断
            final_data=test_data
        else:
            final_data=[]
            for item in test_data:
                if item['case_id'] in eval(mode):
                    final_data.append(item)
        return final_data
if __name__ == '__main__':
    print(GetTestData('data_01-02.xlsx', 'request_data').get_test_data())

                                # 通过传入参数（控制数据的掺入）进而来控制用例的执行
from openpyxl import load_workbook
class GetTestData():
    def __init__(self,filename,sheetname):
        self.filename=filename
        self.sheetname=sheetname
    # 用例的可配置
    def get_test_data(self,mode='all'):
        #通过添加一个参数，可以控制数据的传入情况，进而控制用例的执行（想执行哪条用例就执行哪条用例）
        #mode用来控制是否执行所有用例，默认值为all
        #mode的值只能传入all 和列表两种
        wb=load_workbook(self.filename)
        ws=wb[self.sheetname]
        test_data=[]
        for i in range(1,ws.max_row+1):
            # sub_data={}
            # sub_data['url']=ws.cell(i,ws.max_row).value
            # sub_data['methond']=ws.cell(i,ws.max_row).value
            # sub_data['data']=ws.cell(i,ws.max_row).value
            # sub_data['expected']=ws.cell(i,ws.max_row).value
            # test_data.append(sub_data)

            sub_data= {'case_id':ws.cell(i, 1).value,
                'url': ws.cell(i, 2).value,
                'methond': ws.cell(i, 3).value,
                'data': ws.cell(i,4).value,
                'expected': ws.cell(i, 5).value}
            test_data.append(sub_data)
        if mode=='all':#数据的控制
            final_data=test_data
        else:#[1,2,3,4]
            final_data=[]
            for item in test_data:
                if item['case_id'] in mode:
                    final_data.append(item)
        return final_data
if __name__ == '__main__':
    print(GetTestData('data_01-02.xlsx', 'request_data').get_test_data([1,3]))
