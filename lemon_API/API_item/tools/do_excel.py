# -*- coding: utf-8 -*-
# @Time :2020/6/25 22:25
# @Author : liufei
# @File :do_excel.PY

from openpyxl import load_workbook
from API_item.tools.read_config import ReadConfig
from API_item.tools.project_path import *
from API_item.tools.get_data import GetData

class DoExcel():
    @classmethod
    def get_data(cls,filename):
        tel = getattr(GetData,'NoRegTel')
        test_data=[]
        wb=load_workbook(filename)
        mode=eval(ReadConfig.get_config(test_config_path, 'MODE', 'mode'))
        for key in mode:
            ws = wb[key]
            if mode[key]=='all':
                for i in range(2,ws.max_row+1):
                    raw_data={}
                    raw_data['case_id']=ws.cell(i,1).value
                    raw_data['url']=ws.cell(i,2).value
                    if ws.cell(i,3).value.find('${tel_1}') != -1:
                        raw_data['data']=ws.cell(i,3).value.replace('${tel_1}',str(tel))
                    elif ws.cell(i,3).value.find('${tel}') != -1:
                        raw_data['data'] = ws.cell(i, 3).value.replace('${tel}', str(tel+1))
                    elif ws.cell(i,3).value.find('${admin_tel}') != -1:
                        raw_data['data'] = ws.cell(i, 3).value.replace('${admin_tel}', str(getattr(GetData,'admin_tel')))
                    elif ws.cell(i,3).value.find('${loan_member_id}') != -1:
                        raw_data['data'] = ws.cell(i, 3).value.replace('${loan_member_id}', str(getattr(GetData,'loan_member_id')))
                    elif ws.cell(i,3).value.find('${normal_tel}') != -1:
                        raw_data['data'] = ws.cell(i, 3).value.replace('${normal_tel}', str(getattr(GetData,'normal_tel')))
                    elif ws.cell(i,3).value.find('${normal_tel}') != -1:
                        raw_data['data'] = ws.cell(i, 3).value.replace('${normal_tel}', str(getattr(GetData,'normal_tel')))
                    else:
                        raw_data['data'] = ws.cell(i,3).value
                    if ws.cell(i, 4).value.find('${normal_tel}') != -1:
                        raw_data['check_sql'] = ws.cell(i, 4).value.replace('${normal_tel}', str(getattr(GetData, 'normal_tel')))
                    else:
                        raw_data['check_sql'] = ws.cell(i,4).value
                    raw_data['title'] = ws.cell(i,5).value
                    raw_data['http_method'] = ws.cell(i,6).value
                    raw_data['expected']=ws.cell(i,7).value
                    raw_data['sheet_name'] = key
                    test_data.append(raw_data)
                    cls.update_data(filename, 'init', tel + 2)
            else:
                for case_id in mode[key]:
                    raw_data = {}
                    raw_data['case_id'] = ws.cell(case_id+1, 1).value
                    raw_data['url'] = ws.cell(case_id+1, 2).value
                    if ws.cell(case_id+1, 3).value.find('${tel_1}') != -1:
                        raw_data['data'] = ws.cell(case_id+1, 3).value.replace('${tel_1}', str(tel))
                    elif ws.cell(case_id+1, 3).value.find('${tel}') != -1:
                        raw_data['data'] = ws.cell(case_id+1, 3).value.replace('${tel}', str(tel + 1))
                    elif ws.cell(case_id+1, 3).value.find('${admin_tel}') != -1:
                        raw_data['data'] = ws.cell(case_id+1, 3).value.replace('${admin_tel}',
                                                                       str(getattr(GetData, 'admin_tel')))
                    elif ws.cell(case_id+1, 3).value.find('${loan_member_id}') != -1:
                        raw_data['data'] = ws.cell(case_id+1, 3).value.replace('${loan_member_id}',
                                                                       str(getattr(GetData, 'loan_member_id')))
                    elif ws.cell(case_id+1, 3).value.find('${normal_tel}') != -1:
                        raw_data['data'] = ws.cell(case_id+1, 3).value.replace('${normal_tel}',
                                                                       str(getattr(GetData, 'normal_tel')))
                    elif ws.cell(case_id+1, 3).value.find('${normal_tel}') != -1:
                        raw_data['data'] = ws.cell(case_id+1, 3).value.replace('${normal_tel}',
                                                                       str(getattr(GetData, 'normal_tel')))
                    else:
                        raw_data['data'] = ws.cell(case_id+1,3).value
                    if ws.cell(case_id+1, 4).value.find('${normal_tel}') != -1:
                        raw_data['check_sql'] = ws.cell(case_id+1, 4).value.replace('${normal_tel}', str(getattr(GetData, 'normal_tel')))
                    else:
                        raw_data['check_sql'] = ws.cell(case_id+1,4).value
                    raw_data['title'] = ws.cell(case_id+1, 5).value
                    raw_data['http_method'] = ws.cell(case_id+1, 6).value
                    raw_data['expected'] = ws.cell(case_id+1, 7).value
                    raw_data['sheet_name']=key
                    test_data.append(raw_data)
                    cls.update_data(filename,'init',tel+2)
        return test_data

    @staticmethod
    def write_back(filename,sheetname,row,col,result):#用来将每次的请求结果返回至excel中
        wb = load_workbook(filename)
        ws = wb[sheetname]
        ws.cell(row,col).value=result
        wb.save(filename)

    @classmethod
    def update_data(cls,filename,sheetname,tel):#用来更新excel中的手机号的数据
        wb = load_workbook(filename)
        ws = wb[sheetname]
        ws.cell(2,1).value=tel
        wb.save(filename)



if __name__ == '__main__':
    print(DoExcel.get_data(test_data_path))




