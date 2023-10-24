# -*- coding: utf-8 -*-

"""
一次性获取data目录下所有文件的所有数据
数据存储形式为列表嵌套字典，每一行数据存储为字典。
"""
import os
import sys

curPath = os.path.abspath(os.path.dirname(__file__))
rootPath = os.path.split(curPath)[0]
sys.path.append(rootPath)
import pandas as pd
from conf.path import Path
from common.out_log import OutLog
import os
from openpyxl import load_workbook

logger = OutLog().out_log()


class GetRequestData():
    def get_filename(self):  # 获取目录下所有文件的文件名
        try:
            for root, dirs, files in os.walk(Path.data_thirdApi):
                logger.info("路径{}下的所有文件为{}".format(Path.data_thirdApi, files))
                return files
        except Exception as e:
            logger.error("获取路径{}下的所有文件失败，错误为{}".format(Path.data_thirdApi, e))
            raise e

    def get_request_data(self):
        test_data = []
        filenames = self.get_filename()
        for file in filenames:
            path1 = os.path.split(os.path.split(os.path.realpath(__file__))[0])[0]
            path2 = os.path.join(path1, 'data_thirdApi', file)
            logger.info('开始读取Excel文件{}的测试数据'.format(file))
            sheetnames = pd.ExcelFile(path2)
            logger.info(f'文件{file}中有工作簿{sheetnames.sheet_names}')
            for sheet1 in sheetnames.sheet_names:
                logger.info('正在获取Excel文件{}的{}sheet表单数据'.format(file, sheet1))
                file_data = pd.read_excel(path2, sheet_name=sheet1)
                header_name = file_data.columns.values
                for i in file_data.index.values:
                    row_data = file_data.loc[i, header_name].to_dict()
                    row_data['file_name'] = file
                    row_data['sheet_name'] = sheet1
                    test_data.append(row_data)
        logger.info(f'目录下{Path.data_thirdApi}的所有测试数据已经全部获取完成')
        return test_data

class UpdateData():
    def update(self, file_name, sheet_name, row, column, data):
        wb = load_workbook(file_name)
        ws = wb[sheet_name]
        ws.cell(row, column).value = data
        wb.save(file_name)


if __name__ == '__main__':
    getdata = GetRequestData().get_request_data()
    print(getdata)
