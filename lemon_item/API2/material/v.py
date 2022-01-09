# -*- coding: utf-8 -*-
# @Time :2020/7/18 11:22
# @Author : liufei
# @File :v.PY

from openpyxl import load_workbook
wb=load_workbook(r'D:\Python_files\lemon_item\API\test_data\test_data.xlsx')
ws=wb['register']
print(ws.cell(2,3).value)
print(type(ws.cell(2,3).value))
