# -*- coding: utf-8 -*-
# @Time :2020/7/14 20:25
# @Author : liufei
# @File :do_excel.PY

from openpyxl import load_workbook
from lemon_item.api.common.do_config import DoConfig
from lemon_item.api.common.get_variable import GetVariable
from lemon_item.api.common.get_path import *
from lemon_item.api.common.do_regx import DoRegx

class DoExcel():
    def do_excel(self,filname):
        wb = load_workbook(filname)
        test_data = []
        mode = eval(DoConfig().do_config(case_config_path,'MODE','mode'))
        for key in mode:
            ws = wb[key]
            if mode[key] == 'all':
                for i in range(2, ws.max_row+1):
                    row_data = {}
                    row_data['case_id'] = ws.cell(i, 1).value
                    row_data['url'] = ws.cell(i, 2).value
                    row_data['data'] = DoRegx().do_regx(ws.cell(i, 3).value)
                    row_data['check_sql'] = DoRegx().do_regx(ws.cell(i, 3).value)
                    row_data['title'] = ws.cell(i, 5).value
                    row_data['http_mehod'] = ws.cell(i, 6).value
                    row_data['expected'] = ws.cell(i, 7).value
                    row_data['sheetname'] = key
                    test_data.append(row_data)
                    self.update(filname,'init',getattr(GetVariable, 'NoRegTel')+2)
            else:
                for case_id in mode[key]:
                    row_data = {}
                    row_data['case_id'] = ws.cell(case_id+1, 1).value
                    row_data['url'] = ws.cell(case_id+1, 2).value
                    row_data['data'] = DoRegx().do_regx(ws.cell(case_id+1, 3).value)
                    row_data['check_sql'] = DoRegx().do_regx(ws.cell(case_id+1, 4).value)
                    row_data['title'] = ws.cell(case_id+1, 5).value
                    row_data['http_mehod'] = ws.cell(case_id+1, 6).value
                    row_data['expected'] = ws.cell(case_id + 1, 7).valuerow_data['http_mehod'] = ws.cell(case_id+1, 7).value
                    row_data['sheetname'] = key
                    test_data.append(row_data)
                    self.update(filname, 'init', getattr(GetVariable, 'NoRegTel') + 2)
        return test_data

    def write_back(self, filename, sheetname, row, col, value):
        wb = load_workbook(filename)
        ws = wb[sheetname]
        ws.cell(row, col).value = value
        wb.save(filename)

    def update(self,filename, sheetname, tel):
        wb = load_workbook(filename)
        ws = wb[sheetname]
        ws.cell(2, 1).value = tel
        wb.save(filename)


if __name__ == '__main__':
    print(DoExcel().do_excel(test_data_path))




