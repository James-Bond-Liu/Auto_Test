# -*- coding: utf-8 -*-
# @Time :2020/7/14 20:25
# @Author : liufei
# @File :do_excel.PY

from openpyxl import load_workbook
from API.common.do_config import DoConfig
from API.common.get_variable import GetVariable
from API.common.get_path import *

class DoExcel():
    def do_excel(self,filname):
        wb = load_workbook(filname)
        test_data = []
        mode = eval(DoConfig().do_config(case_config_path,'MODE','mode'))
        for key in mode:
            ws = wb[key]
            if mode[key] == 'all':
                for i in range(2, ws.max_row+1):
                    row_data = {}
                    row_data['case_id'] = ws.cell(i, 1).value
                    row_data['url'] = ws.cell(i, 2).value
                    # if ws.cell(i,3).value.find('${normal_tel}') != -1:
                    #     row_data['data'] = ws.cell(i,3).value.replace('${normal_tel}',str(getattr(GetVariable, 'normal_tel')))
                    # elif ws.cell(i,3).value.find('${NoRegTel}') != -1:
                    #     row_data['data'] = ws.cell(i, 3).value.replace('${NoRegTel}', str(getattr(GetVariable, 'NoRegTel')))
                    # elif ws.cell(i,3).value.find('${NoRegTelR}') != -1:
                    #     row_data['data'] = ws.cell(i, 3).value.replace('${NoRegTelR}', str(3415))
                    # elif ws.cell(i,3).value.find('${admin_tel}') != -1:
                    #     row_data['data'] = ws.cell(i, 3).value.replace('${admin_tel}', str(getattr(GetVariable, 'admin_tel')))
                    # elif ws.cell(i,3).value.find('${loan_member_id}') != -1:
                    #     row_data['data'] = ws.cell(i, 3).value.replace('${loan_member_id}', str(getattr(GetVariable, 'loan_member_id')))
                    # elif ws.cell(i,3).value.find('${memberId}') != -1:
                    #     row_data['data'] = ws.cell(i, 3).value.replace('${memberId}', str(getattr(GetVariable, 'member_id')))
                    # else:
                    row_data['data'] = ws.cell(i, 3).value
                    row_data['title'] = ws.cell(i, 4).value
                    row_data['http_mehod'] = ws.cell(i, 5).value
                    row_data['expected'] = ws.cell(i, 6).value
                    test_data.append(row_data)
                    self.update(filname,'init',getattr(GetVariable, 'NoRegTel')+2)
            else:
                for case_id in mode[key]:
                    row_data = {}
                    row_data['case_id'] = ws.cell(case_id+1, 1).value
                    row_data['url'] = ws.cell(case_id+1, 2).value
                    # if ws.cell(case_id+1, 3).value.find('${normal_tel}') != -1:
                    #     row_data['data'] = ws.cell(case_id+1, 3).value.replace('${normal_tel}',str(getattr(GetVariable,'normal_tel')))
                    # elif ws.cell(case_id+1, 3).value.find('${NoRegTel}') != -1:
                    #     row_data['data'] = ws.cell(case_id+1, 3).value.replace('${NoRegTel}', str(getattr(GetVariable, 'NoRegTel')))
                    # elif ws.cell(case_id+1, 3).value.find('${admin_tel}') != -1:
                    #     row_data['data'] = ws.cell(case_id+1, 3).value.replace('${admin_tel}', str(getattr(GetVariable, 'admin_tel')))
                    # elif ws.cell(case_id+1, 3).value.find('${loan_member_id}') != -1:
                    #     row_data['data'] = ws.cell(case_id+1, 3).value.replace('${loan_member_id}', str(getattr(GetVariable, 'loan_member_id')))
                    # elif ws.cell(case_id+1, 3).value.find('${memberId}') != -1:
                    #     row_data['data'] = ws.cell(case_id+1, 3).value.replace('${memberId}', str(getattr(GetVariable, 'memeberId')))
                    # else:
                    row_data['data'] = ws.cell(case_id+1, 3).value
                    row_data['title'] = ws.cell(case_id+1, 4).value
                    row_data['http_mehod'] = ws.cell(case_id+1, 5).value
                    row_data['expected'] = ws.cell(case_id+1, 6).value
                    test_data.append(row_data)
                    self.update(filname, 'init', getattr(GetVariable, 'NoRegTel') + 2)
        for item in test_data:
            if item['data'].find('${normal_tel}') != -1:
                item['data'] = item['data'].replace('${normal_tel}',str(getattr(GetVariable, 'normal_tel')))
            elif item['data'].find("${NoRegTel}") != -1:
                item['data'] = item['data'].replace("${NoRegTel}",str(getattr(GetVariable, 'NoRegTel')))
            elif item['data'].find('${NoRegTelR}') != -1:
                item['data'] = item['data'].replace('${NoRegTelR}',str(getattr(GetVariable, 'NoRegTelR')))
            elif item['data'].find('${admin_tel}') != -1:
                item['data'] = item['data'].replace('${admin_tel}',str(getattr(GetVariable, 'admin_tel')))
            elif item['data'].find('${loan_member_id}') != -1:
                item['data'] = item['data'].replace('${loan_member_id}',str(getattr(GetVariable, 'loan_member_id')))
            elif item['data'].find('${memberId}') != -1:
                item['data'] = item['data'].replace('${member_id}',str(getattr(GetVariable, 'member_id')))
            else:
                pass

        return test_data
    def write_back(self, filename, sheetname, row, col, value):
        wb=load_workbook(filename)
        ws=wb[sheetname]
        ws.cell(row, col).value=value
        wb.save(filename)

    def update(self, filename, sheetname, tel):
        wb=load_workbook(filename)
        ws=wb[sheetname]
        ws.cell(2, 1).value = tel
        wb.save(filename)


if __name__ == '__main__':
    print(DoExcel().do_excel(test_data_path))



